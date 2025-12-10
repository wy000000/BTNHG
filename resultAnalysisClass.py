from BTNHGV2ParameterClass import BTNHGV2ParameterClass
from ExtendedNNModule import ExtendedNNModule
import os
import datetime
from dataclasses import dataclass, asdict
import numpy as np
import shutil
import torch
import torch.nn as nn
import matplotlib.pyplot as plt
import platform
import psutil
import sys
import cpuinfo
import torch.nn.functional as F
import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, Reference
from sklearn.metrics import (
	accuracy_score,
	balanced_accuracy_score,
	precision_score,
	recall_score,
	f1_score,
	roc_auc_score,
	average_precision_score,
	cohen_kappa_score,
	matthews_corrcoef,
	confusion_matrix,
	ConfusionMatrixDisplay)

class resultAnalysisClass:
	def __init__(self,
			  	model:ExtendedNNModule=None,
				folderPath:str=BTNHGV2ParameterClass.dataPath,
				resultFolderName:str=BTNHGV2ParameterClass.resultFolderName				
				):
		#path
		self.resultFolderName=resultFolderName
		self.resultFolderPath=os.path.join(folderPath, self.resultFolderName)
		self.methodFolderName=None
		self.methodFolderPath=None
		self.model=model
		self.modelName=self.model.__class__.__name__
		self.model.modelName=self.modelName
		#评估参数
		self.accuracy=0.0
		self.model.evaluationMetrics=self.compute_metrics()
		#配置信息
		self.model.env_info=self.get_training_env_info()	

	def showEvaluationMetrics(self):
		metrics=self.model.evaluationMetrics
		if(metrics is None):
			self.model.evaluationMetrics=self._compute_metrics()
			metrics=self.model.evaluationMetrics
		if(metrics is None):
			print("Evaluation metrics are None.")
			return None
		#按行打印metrics
		for key, value in metrics.items():
			print(f"{key}: {value}")
		return	

	def compute_metrics(self):

		# 从模型中获取真实标签、预测标签和概率
		y_true = self.model.all_y_true
		y_preds = self.model.all_preds
		y_probs = self.model.all_probs

		if(y_true is None or y_preds is None or y_probs is None):
			return None

		# 转换为 numpy
		y_true = y_true.cpu().numpy() if isinstance(y_true, torch.Tensor) else np.array(y_true)
		y_preds = y_preds.cpu().numpy() if isinstance(y_preds, torch.Tensor) else np.array(y_preds)
		y_probs = y_probs.cpu().numpy() if isinstance(y_probs, torch.Tensor) else np.array(y_probs)

		# 基础指标
		acc = accuracy_score(y_true, y_preds)
		self.accuracy=acc
		self.model.accuracy=acc
		bal_acc = balanced_accuracy_score(y_true, y_preds)

		# Precision / Recall / F1
		prec_macro = precision_score(y_true, y_preds, average="macro", zero_division=0)
		rec_macro = recall_score(y_true, y_preds, average="macro", zero_division=0)
		f1_macro = f1_score(y_true, y_preds, average="macro", zero_division=0)

		prec_weighted = precision_score(y_true, y_preds, average="weighted", zero_division=0)
		rec_weighted = recall_score(y_true, y_preds, average="weighted", zero_division=0)
		f1_weighted = f1_score(y_true, y_preds, average="weighted", zero_division=0)

		# ROC-AUC & PR-AUC (多分类时用宏平均)
		try:
			roc_auc = roc_auc_score(y_true, y_probs, multi_class="ovr", average="macro")
		except Exception:
			roc_auc = None

		try:
			pr_auc = average_precision_score(y_true, y_probs, average="macro")
		except Exception:
			pr_auc = None

		# Cohen's Kappa & MCC
		kappa = cohen_kappa_score(y_true, y_preds)
		mcc = matthews_corrcoef(y_true, y_preds)

		# 平均置信度
		confidences = y_probs.max(axis=1)
		avg_conf = confidences.mean()

		metrics={			
			"accuracy": acc,
			"training_time": self.model.training_time,
			"avg_confidence": avg_conf,
			"balanced_accuracy": bal_acc,
			"precision_macro": prec_macro,
			"recall_macro": rec_macro,
			"f1_macro": f1_macro,
			"precision_weighted": prec_weighted,
			"recall_weighted": rec_weighted,
			"f1_weighted": f1_weighted,
			"roc_auc_macro": roc_auc,
			"pr_auc_macro": pr_auc,
			"cohen_kappa": kappa,
			"mcc": mcc
		}
		# self.evaluationMetrics=metrics
		print("Evaluation metrics are computed.")
		return metrics

	def plot_true_pred_counts(self, y_true=None, y_preds=None):
		if(y_true is None):
			y_true=self.model.all_y_true
		if(y_preds is None):
			y_preds=self.model.all_preds
		if(y_true is None or y_preds is None):
			print("y_true or y_preds is None.")
			return None
		
		# 绘制预测与真实数量的柱状图
		# 获取所有类别
		classes = np.unique(np.concatenate([y_preds, y_true]))

		# 统计每个类别的预测数量和真实数量
		pred_counts = [np.sum(np.array(y_preds) == c) for c in classes]
		label_counts = [np.sum(np.array(y_true) == c) for c in classes]

		# 绘制柱状图
		x = np.arange(len(classes))  # 类别索引
		width = 0.35  # 柱子宽度

		fig, ax = plt.subplots(figsize=(8, 6))
		rects1 = ax.bar(x - width/2, pred_counts, width, label='Predicted', color='skyblue')
		rects2 = ax.bar(x + width/2, label_counts, width, label='True', color='salmon')

		# 添加标签和标题
		ax.set_xlabel('Classes')
		ax.set_ylabel('Counts')
		ax.set_title('Predicted vs True Counts per Class')
		ax.set_xticks(x)
		ax.set_xticklabels(classes)
		ax.legend()

		# 在柱子上标注数值
		for rects in [rects1, rects2]:
			for rect in rects:
				height = rect.get_height()
				ax.annotate(f'{height}',
							xy=(rect.get_x() + rect.get_width() / 2, height),
							xytext=(0, 3),  # 向上偏移 3
							textcoords="offset points",
							ha='center', va='bottom')

		plt.tight_layout()
		plt.show()
		return
	
	def plot_confusion_matrix(self, y_true=None, y_preds=None):
		if(y_true is None):
			y_true=self.model.all_y_true
		if(y_preds is None):
			y_preds=self.model.all_preds
		if(y_true is None or y_preds is None):
			print("y_true or y_preds is None.")
			return None
		cm = confusion_matrix(y_true, y_preds)
		disp = ConfusionMatrixDisplay(confusion_matrix=cm)
		disp.plot(cmap=plt.cm.Blues)
		plt.title("Confusion Matrix on Test Set")
		plt.show()

	def get_training_env_info(self):
		"""
		获取神经网络训练常用的软件和硬件配置信息
		"""
		env_info = {}

		# 🖥️ 硬件信息
		env_info["System"] = platform.system()
		env_info["Architecture"] = cpuinfo.get_cpu_info()['arch']
		env_info["CPU"] = cpuinfo.get_cpu_info()['brand_raw']
		env_info["Cores"] = psutil.cpu_count(logical=False)
		env_info["memory_total_GB"] = round(psutil.virtual_memory().total / (1024**3), 2)

		# GPU 信息（PyTorch）
		if torch.cuda.is_available():
			env_info["gpu_available"] = True
			env_info["gpu_count"] = torch.cuda.device_count()

			for i in range(env_info["gpu_count"]):
				props = torch.cuda.get_device_properties(i)
				name = torch.cuda.get_device_name(i)
				total_mem = round(props.total_memory / (1024**3), 2)  # 转换为 GB
				env_info[f"GPU {i}"] = f"{name} {total_mem} GB"
		else:
			env_info["gpu_available"] = False

		# 📦 软件信息
		env_info["python_version"] = sys.version.split()[0]
		env_info["torch_version"] = torch.__version__
		return env_info

	def showExtendedAttributes(self):
		print(self.model.serialize_extended_attributes())

	def save(self,
			save:bool=BTNHGV2ParameterClass.save,
			saveModelStateDict:bool=BTNHGV2ParameterClass.saveModelStateDict,
			saveFullModel:bool=BTNHGV2ParameterClass.saveFullModel):
		'''
		Items can be saved only if save=True
		'''
		if save:
			self._createMethodFolder(kfold=False)
			self._saveBTNHGV2ParameterClass()
			self._saveExtendedAttributes()
			self._save_epoch_loss_list()
			self._saveY_true_preds_probs()
			
			if(saveModelStateDict):
				self._saveModel_state_dict()
			if(saveFullModel):
				self._saveFullModel()
		return
	
	def save_kFold(self,
			save:bool=BTNHGV2ParameterClass.save,
			saveModelStateDict:bool=BTNHGV2ParameterClass.saveModelStateDict,
			saveFullModel:bool=BTNHGV2ParameterClass.saveFullModel):
		'''
		Items can be saved only if save=True
		'''
		if save:
			self._createMethodFolder(kfold=True)
			self._saveBTNHGV2ParameterClass()
			self._save_extended_attributes_kFold()
			self._save_evaluationMetrics_kFold()
			if(saveModelStateDict):
				self._save_kFold_best_model_state_dict()
			if(saveFullModel):
				self._save_kFold_fullModel()
			
		return
	
	def _createMethodFolder(self, kfold:bool=False):
		modelName=self.modelName
		accuracy=self.accuracy
		#folderName=modelName+年月日时分秒+accuracy,以"-"分隔
		if kfold:
			self.methodFolderName=modelName+"-kFold-"+datetime.datetime.now().strftime("%Y.%m.%d %H.%M.%S")\
					+"-"+f"acc {accuracy:.4f}"
		else:
			self.methodFolderName=modelName+"-"+datetime.datetime.now().strftime("%Y.%m.%d %H.%M.%S")\
					+"-"+f"acc {accuracy:.4f}"
		#将path与folderName拼接
		self.methodFolderPath=os.path.join(self.resultFolderPath, self.methodFolderName)
		print(f"methodFolderPath={self.methodFolderPath}")
		#创建文件夹
		os.makedirs(self.methodFolderPath, exist_ok=True)

		return self.methodFolderPath
	
	def _saveBTNHGV2ParameterClass(self):
		#复制BTNHGV2ParameterClass.py 到 folderPath
		fileName=BTNHGV2ParameterClass.__name__+".txt"
		filePath=os.path.join(self.methodFolderPath, fileName)

		shutil.copyfile(fileName, filePath)
		print(f"{fileName}已保存")

		return filePath
	
	def _saveExtendedAttributes(self):
		fileName=BTNHGV2ParameterClass.extendedAttributesFileName
		filePath=os.path.join(self.methodFolderPath, fileName)
		str=self.model.serialize_extended_attributes()
		
		with open(filePath, "w") as f:
			f.write(str)
		print(f"{fileName}已保存")
		
		return filePath

	def _save_epoch_loss_list(self):
		fileName = BTNHGV2ParameterClass.epoch_loss_listFileName
		filePath = os.path.join(self.methodFolderPath, fileName)

		epoch_loss_list = []
		if self.model.epoch_loss_list is not None:
			epoch_loss_list = self.model.epoch_loss_list

		# 创建 DataFrame
		df = pd.DataFrame(epoch_loss_list, columns=['Epoch', 'Loss'])

		# 保存到 Excel，并生成散点图
		with pd.ExcelWriter(filePath, engine="xlsxwriter") as writer:
			df.to_excel(writer, sheet_name="Sheet1", index=False)

			workbook  = writer.book
			worksheet = writer.sheets["Sheet1"]

			# 创建散点图（带平滑线）
			chart = workbook.add_chart({"type": "scatter", "subtype": "smooth"})

			chart.add_series({
				"name":       "Loss Curve",
				"categories": ["Sheet1", 1, 0, len(df), 0],  # Epoch 列
				"values":     ["Sheet1", 1, 1, len(df), 1],  # Loss 列
				"marker":     {"type": "circle"},
				"smooth":     True
			})

			chart.set_title({"name": "Epoch vs Loss"})
			chart.set_x_axis({"name": "Epoch"})
			chart.set_y_axis({"name": "Loss"})

			# 插入图表到 Excel
			worksheet.insert_chart("D2", chart)

		print(f"{fileName}已保存")
		return filePath

	def _saveY_true_preds_probs(self):
		fileName = BTNHGV2ParameterClass.y_true_preds_probsFileName
		filePath = os.path.join(self.methodFolderPath, fileName)

		y_true = self.model.all_y_true
		y_preds = self.model.all_preds
		y_probs = self.model.all_probs

		# print(f"y_true.shape={getattr(y_true, 'shape', None)}")
		# print(f"y_preds.shape={getattr(y_preds, 'shape', None)}")
		# print(f"y_probs.shape={getattr(y_probs, 'shape', None)}")
	
		# 如果 y_probs 是二维数组/张量，拆成多列
		df_probs = pd.DataFrame(y_probs, columns=[f"prob_{i}" for i in range(len(y_probs[0]))]) \
				if y_probs is not None and hasattr(y_probs[0], "__len__") else pd.DataFrame({"prob": y_probs})

		# 拼接成完整 DataFrame
		df = pd.DataFrame({
			"y_true": y_true,
			"y_preds": y_preds
		})
		df = pd.concat([df, df_probs], axis=1)

		df.to_excel(filePath, index=False)
		print(f"{fileName}已保存")
		return filePath

	def _saveModel_state_dict(self):
		fileName=BTNHGV2ParameterClass.modelStateDictFileName
		filePath=os.path.join(self.methodFolderPath, fileName)

		torch.save(self.model.state_dict(), filePath)
		print(f"{fileName}已保存")

		return filePath
	
	def _saveFullModel(self):
		#复制model 到 self.methodFolderPath
		fileName=BTNHGV2ParameterClass.fullModelFileName
		filePath=os.path.join(self.methodFolderPath, fileName)

		torch.save(self.model, filePath)
		print(f"{fileName}已保存")

		return filePath
		
	def _save_kFold_best_model_state_dict(self):
		fileName=BTNHGV2ParameterClass.modelStateDictFileName
		filePath=os.path.join(self.methodFolderPath, fileName)

		torch.save(self.model.kFold_best_model_state, filePath)
		print(f"{fileName}已保存")

		return filePath

	def _save_kFold_fullModel(self):
		#复制model 到 self.methodFolderPath
		fileName=BTNHGV2ParameterClass.fullModelFileName
		filePath=os.path.join(self.methodFolderPath, fileName)
		self.model.load_state_dict(self.model.kFold_best_model_state)

		torch.save(self.model, filePath)
		print(f"{fileName}已保存")

		return filePath
	
	def _save_extended_attributes_kFold(self):
		fileName=BTNHGV2ParameterClass.extendedAttributesFileName
		filePath=os.path.join(self.methodFolderPath, fileName)
		str=self.model.serialize_extended_attributes_kFold()
		
		with open(filePath, "w") as f:
			f.write(str)
		print(f"{fileName}已保存")
		
		return filePath
	
	def _save_evaluationMetrics_kFold(self, filename="evaluation_metrics.xlsx"):
		fileName=BTNHGV2ParameterClass.evaluationMetricsFileName
		filePath=os.path.join(self.methodFolderPath, fileName)
		kFold_evaluations=self.model.kFold_evaluations
		# 创建工作簿
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Metrics"

		# 获取所有指标名（可能不同 fold 有不同指标）
		all_metrics = set()
		for eval_dict in kFold_evaluations:
			all_metrics.update(eval_dict.keys())
		all_metrics = list(all_metrics)

		# 写标题行
		headers = ["Metric"] + [f"Fold{i+1}" for i in range(len(kFold_evaluations))] + ["Average"]
		ws.append(headers)

		# 写数据行
		for metric in all_metrics:
			row = [metric]
			for i, eval_dict in enumerate(kFold_evaluations):
				value = eval_dict.get(metric, None)
				row.append(value)
			# 添加平均值公式
			start_col = 2
			end_col = 1 + len(kFold_evaluations)
			avg_formula = f"=AVERAGE({ws.cell(row=ws.max_row+1, column=start_col).coordinate}\
				:{ws.cell(row=ws.max_row+1, column=end_col).coordinate})"
			row.append(avg_formula)
			ws.append(row)

		# 绘制 accuracy 折线图
		chart = LineChart()
		chart.title = "Accuracy across folds"
		chart.y_axis.title = "Accuracy"
		chart.x_axis.title = "Fold"

		# 找到 accuracy 行
		for row_idx in range(2, ws.max_row + 1):
			if ws.cell(row=row_idx, column=1).value == "accuracy":
				data = Reference(ws, min_col=2, max_col=1+len(kFold_evaluations), min_row=row_idx, max_row=row_idx)
				chart.addData(data, titles_from_data=False)
				cats = Reference(ws, min_col=2, max_col=1+len(kFold_evaluations), min_row=1, max_row=1)
				chart.set_categories(cats)
				ws.add_chart(chart, f"E{row_idx+2}")
				break

		# 保存文件
		wb.save(filePath)
		print(f"{fileName}已保存")
		return filePath

		


	

